#!/usr/bin/env python3
"""
Monsters Among Us (MAU) Podcast Scraper & Data Pipeline v2
============================================================
Scrapes every episode of the MAU podcast via its PUBLIC RSS feed,
transcribes audio locally using OpenAI Whisper (free, no API key),
parses caller stories via LLM (Anthropic Claude or OpenAI), and
outputs tidy data.

NO SPOTIFY DEVELOPER ACCOUNT NEEDED.

Requirements:
    pip install feedparser requests openpyxl anthropic openai tqdm python-dotenv openai-whisper nltk

API Provider:
    Choose at runtime with --api flag (no file editing needed):
      --api haiku          Claude Haiku (default, cheapest)
      --api sonnet         Claude Sonnet (best quality)
      --api o4-mini        OpenAI o4-mini
      --api gpt-4o-mini    OpenAI GPT-4o mini (very cheap)
      --api gpt-4.1-mini   OpenAI GPT-4.1 mini

Usage:
    1. Configure .env file with API key(s) (see .env.example)
    2. python mau_scraper_v2.py --step demo                    # Demo
    3. python mau_scraper_v2.py --step all --api haiku         # Full pipeline with Haiku
    4. python mau_scraper_v2.py --step parse --api o4-mini     # Parse with OpenAI
    5. python mau_scraper_v2.py --step transcribe --limit 5    # Transcribe 5 episodes
    6. python mau_scraper_v2.py --step export                  # Export to XLSX
    7. python mau_scraper_v2.py --step reparse-regions --api haiku  # Fix locations only
    8. python mau_scraper_v2.py --step reset-parsed            # Wipe parsed JSON only
    9. python mau_scraper_v2.py --step reset                   # Wipe ALL cached data

Notes:
    - Before any work begins, the script prints a time/cost estimate showing
      how many episodes need processing vs. how many are already cached.
    - During parsing, a live ETA updates after each episode completes.
    - Each parsed call is verified against the original transcript to flag
      potential hallucinations. The 'verified' column in the XLSX shows the result.
    - Export auto-versions: MAU_Tidy_Data.xlsx → _v2.xlsx → _v3.xlsx, etc.
    - reset-parsed is useful when switching API providers or changing the prompt.
    - reset wipes everything for a fully clean run from scratch.
"""

import os
import re
import json
import time
import logging
import argparse
from pathlib import Path
from typing import Optional
from datetime import datetime, timedelta

import feedparser
import requests
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────
load_dotenv()  # Loads .env from current working directory (or parent dirs)

# ── API PROVIDER — auto-detected from model name via --api flag ──────────
# No need to edit this. Use --api on the command line:
#   --api haiku          (Claude Haiku — cheapest)
#   --api sonnet         (Claude Sonnet — best quality)
#   --api o4-mini        (OpenAI o4-mini)
#   --api gpt-4o-mini    (OpenAI GPT-4o mini — very cheap)
#   --api gpt-4.1-mini   (OpenAI GPT-4.1 mini)

MODEL_REGISTRY = {
    "haiku":         {"model": "claude-haiku-4-5-20251001",  "provider": "anthropic", "input": 1.00, "output": 5.00},
    "sonnet":        {"model": "claude-sonnet-4-20250514",   "provider": "anthropic", "input": 3.00, "output": 15.00},
    "o4-mini":       {"model": "o4-mini-2025-04-16",         "provider": "openai",    "input": 1.10, "output": 4.40},
    "gpt-4o-mini":   {"model": "gpt-4o-mini",               "provider": "openai",    "input": 0.15, "output": 0.60},
    "gpt-4.1-mini":  {"model": "gpt-4.1-mini",              "provider": "openai",    "input": 0.40, "output": 1.60},
}
DEFAULT_API = "haiku"

# MAU RSS Feeds (public, no auth needed)
# The show has migrated hosts over the years; Audioboom is current.
RSS_FEEDS = [
    "https://feeds.audioboom.com/channels/5106512.rss",         # Audioboom (current)
    "http://monstersamonguspodcast.libsyn.com/rss",             # Libsyn (archive)
    "https://feeds.megaphone.fm/QCD6797102910",                 # Megaphone (archive)
]

# Whisper model size — "base" is fast, "medium" is recommended, "large" is most accurate
WHISPER_MODEL = os.getenv("WHISPER_MODEL", "medium")

# Directories
BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
AUDIO_DIR = DATA_DIR / "audio"
TRANSCRIPTS_DIR = DATA_DIR / "transcripts"
PARSED_DIR = DATA_DIR / "parsed"
OUTPUT_DIR = DATA_DIR / "output"

for d in [DATA_DIR, AUDIO_DIR, TRANSCRIPTS_DIR, PARSED_DIR, OUTPUT_DIR]:
    d.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        # ADD encoding="utf-8" HERE:
        logging.FileHandler(DATA_DIR / "scraper.log", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)

log = logging.getLogger("mau_scraper")


# ─────────────────────────────────────────────────────────────────────────────
# ENTITY TAXONOMY
# ─────────────────────────────────────────────────────────────────────────────
ENTITY_TAXONOMY = """
CRYPTID ENTITIES:
  Bigfoot/Sasquatch, Dogman, Mothman, Skinwalker, Wendigo, Thunderbird,
  Crawler/Rake, Chupacabra, Mutant Canine/Turner Beast, Lake Monster,
  Black-Eyed Kids (BEK), Not Deer, Goatman, Jersey Devil, Unknown Cryptid

GHOSTLY/SPIRITUAL:
  Ghost/Apparition, Shadow Person, Hat Man, Old Hag, Poltergeist,
  Demonic Entity, Angel/Positive Spirit, Doppelganger, Residual Haunting,
  Intelligent Haunting, Portal/Vortex, Afterlife Communication (ADC),
  Sleep Paralysis Entity

UFO/AERIAL:
  UFO/UAP, Black Triangle, Orb/Light, Fireball, Phantom Lights,
  Abduction, Alien Entity, Satellite/Rocket (Explained)

HIGH STRANGENESS:
  Missing Time, Glitch in the Matrix, Time Slip, Premonition/Precognition,
  Telepathy, Psychic Experience, Astral Projection, Aura/Energy,
  Synchronicity, Men in Black, Phantom Person, Vanishing Object,
  Dime/Coin from the Dead

ENVIRONMENTAL/SENSORY:
  Phantom Sound, Phantom Smell, Unexplained Animal Behavior,
  Electronic Malfunction, Temperature Anomaly

OTHER:
  Coincidence, Unidentified Animal, Explained (mundane), Dream/Vision,
  Curse/Hex, Fairy/Fae, Humanoid (unclassified), Other
"""


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1: FETCH EPISODE LIST FROM RSS (NO API KEY NEEDED)
# ─────────────────────────────────────────────────────────────────────────────
def fetch_episode_list() -> list[dict]:
    """
    Parses the MAU RSS feed(s) to get all episodes with metadata and audio URLs.
    RSS feeds are public — no authentication required.
    """
    cache_path = DATA_DIR / "episode_list.json"
    if cache_path.exists():
        log.info("Loading cached episode list...")
        with open(cache_path) as f:
            return json.load(f)

    log.info("Fetching episode list from RSS feed(s)...")
    seen_titles = set()
    episodes = []

    for feed_url in RSS_FEEDS:
        log.info(f"  Parsing: {feed_url[:60]}...")
        try:
            feed = feedparser.parse(feed_url)
            for entry in feed.entries:
                title = entry.get("title", "")
                if title in seen_titles:
                    continue
                seen_titles.add(title)

                # Extract audio URL from enclosures
                audio_url = None
                for enc in entry.get("enclosures", []):
                    if enc.get("type", "").startswith("audio/"):
                        audio_url = enc.get("href") or enc.get("url")
                        break
                if not audio_url:
                    links = entry.get("links", [])
                    for link in links:
                        if link.get("type", "").startswith("audio/"):
                            audio_url = link.get("href")
                            break

                # Parse season/episode from title
                season_num, ep_num = _parse_season_episode(title)

                # Publication date
                pub_date = entry.get("published", "")
                # Normalize to YYYY-MM-DD
                if hasattr(entry, "published_parsed") and entry.published_parsed:
                    import calendar
                    t = entry.published_parsed
                    pub_date = f"{t.tm_year}-{t.tm_mon:02d}-{t.tm_mday:02d}"

                # Duration
                duration = entry.get("itunes_duration", "")

                episode_id = re.sub(r"[^\w]", "_", title)[:60]

                episodes.append({
                    "episode_id": episode_id,
                    "name": title,
                    "release_date": pub_date,
                    "duration": duration,
                    "description": entry.get("summary", "")[:500],
                    "audio_url": audio_url,
                    "season": season_num,
                    "episode": ep_num,
                    "feed_source": feed_url[:40],
                })
        except Exception as e:
            log.warning(f"  Failed to parse {feed_url}: {e}")

    episodes.sort(key=lambda e: e.get("release_date", ""))

    with open(cache_path, "w") as f:
        json.dump(episodes, f, indent=2)

    log.info(f"Total unique episodes found: {len(episodes)}")
    return episodes


def _parse_season_episode(title: str) -> tuple[Optional[int], Optional[int]]:
    patterns = [
        r"[Ss](?:n\.?\s*|eason\s*)(\d+)\s*[Ee]p(?:isode)?\.?\s*(\d+)",
        r"[Ss](\d+)\s*[Ee]p?(\d+)",
        r"\(Sn\.\s*(\d+)\s*Ep\.\s*(\d+)\)",
    ]
    for pat in patterns:
        m = re.search(pat, title)
        if m:
            return int(m.group(1)), int(m.group(2))
    m = re.search(r"[Ee]p(?:isode)?\.?\s*(\d+)", title)
    if m:
        return None, int(m.group(1))
    return None, None


def _clean_audio_url(url: str) -> str:
    """
    Strips podcast tracking/analytics redirect wrappers to get the direct MP3 URL.

    Many podcast hosts chain URLs through trackers like arttrk.com, podtrac.com,
    pscrb.fm, clrtpod.com, etc. These can fail when DNS doesn't resolve.
    The actual audio lives on audioboom.com (or similar CDN).

    Example input:
      https://arttrk.com/p/ABMA5/clrtpod.com/m/pscrb.fm/rss/p/dts.podtrac.com/
      redirect.mp3/audioboom.com/posts/8633435.mp3?modified=1738965670&sid=...

    Output:
      https://audioboom.com/posts/8633435.mp3
    """
    # Look for audioboom.com URL buried in the redirect chain
    m = re.search(r"(audioboom\.com/posts/\d+\.mp3)", url)
    if m:
        clean = f"https://{m.group(1)}"
        if clean != url:
            log.debug(f"  URL cleaned: {url[:60]}... → {clean}")
        return clean

    # Look for other common CDN hosts buried in tracker chains
    for host in ["megaphone.fm", "libsyn.com", "podbean.com", "buzzsprout.com"]:
        m = re.search(rf"((?:https?://)?[^/]*{re.escape(host)}/[^\s?]+\.mp3)", url)
        if m:
            clean = m.group(1)
            if not clean.startswith("http"):
                clean = f"https://{clean}"
            return clean

    # No tracker detected — return as-is
    return url


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2: DOWNLOAD AUDIO & TRANSCRIBE WITH WHISPER (FREE, LOCAL)
# ─────────────────────────────────────────────────────────────────────────────
def fetch_transcripts(episodes: list[dict]) -> None:
    """
    For each episode:
      1. Download the MP3 from the RSS audio_url
      2. Transcribe locally using OpenAI Whisper (free, runs on your CPU/GPU)
      3. Save transcript with timestamps as JSON + plain text

    Whisper produces word-level timestamps, which we use for call_start_time.
    GPU acceleration available via CUDA (NVIDIA) or ROCm (AMD) if configured.
    """
    try:
        import whisper
    except ImportError:
        log.error(
            "Whisper not installed. Run: pip install openai-whisper\n"
            "Also needs ffmpeg: sudo apt install ffmpeg (Linux) or choco install ffmpeg (Windows)"
        )
        raise

    log.info(f"Loading Whisper model '{WHISPER_MODEL}'...")
    model = whisper.load_model(WHISPER_MODEL)

    for ep in tqdm(episodes, desc="Transcribing"):
        eid = ep["episode_id"]
        transcript_path = TRANSCRIPTS_DIR / f"{eid}.txt"
        timestamps_path = TRANSCRIPTS_DIR / f"{eid}_timestamps.json"

        if transcript_path.exists() and transcript_path.stat().st_size > 500:
            continue

        audio_url = ep.get("audio_url")
        if not audio_url:
            log.warning(f"  No audio URL: {ep['name'][:60]}")
            with open(transcript_path, "w") as f:
                f.write(f"TRANSCRIPT_UNAVAILABLE\n{ep['name']}")
            continue

        # Strip tracking redirects (arttrk.com, podtrac, etc.) to get the real MP3 URL
        audio_url = _clean_audio_url(audio_url)

        # Download audio
        audio_path = AUDIO_DIR / f"{eid}.mp3"
        if not audio_path.exists() or audio_path.stat().st_size < 100000:
            log.info(f"  Downloading: {ep['name'][:50]}...")
            tmp_path = audio_path.with_suffix(".mp3.tmp")
            try:
                resp = requests.get(audio_url, stream=True, timeout=120)
                resp.raise_for_status()
                with open(tmp_path, "wb") as f:
                    for chunk in resp.iter_content(chunk_size=8192):
                        f.write(chunk)
                tmp_path.replace(audio_path)
            except Exception as e:
                log.error(f"  Download failed: {e}")
                tmp_path.unlink(missing_ok=True)
                continue

        # Transcribe with Whisper
        log.info(f"  Transcribing: {ep['name'][:50]}...")
        try:
            result = model.transcribe(
                str(audio_path),
                language="en",
                verbose=False,
            )

            # Save plain text transcript (atomic: write tmp then rename)
            tmp_txt = transcript_path.with_suffix(".txt.tmp")
            with open(tmp_txt, "w", encoding="utf-8") as f:
                f.write(result["text"])

            # Save timestamped segments for call_start_time lookup
            segments = []
            for seg in result.get("segments", []):
                segments.append({
                    "start_ms": int(seg["start"] * 1000),
                    "start_hms": _seconds_to_hms(seg["start"]),
                    "end_hms": _seconds_to_hms(seg["end"]),
                    "text": seg["text"].strip(),
                })
            tmp_ts = timestamps_path.with_suffix(".json.tmp")
            with open(tmp_ts, "w", encoding="utf-8") as f:
                json.dump(segments, f, indent=2)

            # Both files written successfully — commit by renaming
            tmp_txt.replace(transcript_path)
            tmp_ts.replace(timestamps_path)

            log.info(f"  ✓ Transcribed: {ep['name'][:50]}")

        except Exception as e:
            log.error(f"  Whisper failed: {e}")
            # Clean up any partial temp files
            for tmp in [transcript_path.with_suffix(".txt.tmp"),
                        timestamps_path.with_suffix(".json.tmp")]:
                tmp.unlink(missing_ok=True)


def _seconds_to_hms(seconds: float) -> str:
    total = int(seconds)
    h = total // 3600
    m = (total % 3600) // 60
    s = total % 60
    return f"{h}:{m:02d}:{s:02d}"


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3: PARSE TRANSCRIPTS (CLAUDE-POWERED — same as v1)
# ─────────────────────────────────────────────────────────────────────────────
PARSE_SYSTEM_PROMPT = f"""You are a structured data extraction system. Your output will be used directly
for statistical analysis in R and Python. Human readability is secondary to
data consistency and type safety.

Output valid JSON only. Do not include explanations, comments, or uncertainty
markers. All fields must be present for every record. Unknown or missing values
must be returned as JSON null. NEVER return strings such as "Unknown", "N/A",
"Unclear", "Parts Unknown", or empty strings — use null instead.

You are analyzing transcripts of the podcast "Monsters Among Us" (MAU).
Identify EVERY individual caller story and extract one JSON object per caller.

CALLER BOUNDARY RULES:
- A caller is identified by Derek's introduction, the caller's self-introduction,
  or a clear topic/story transition.
- Derek Hayes is the HOST — never create a record for him.
- Advertisements, promos, and show metadata are NOT calls. Skip them.
- If a caller calls back (continuation), merge into ONE record unless the
  topic changes substantially.
- "Beyond" / Patreon / "sleep paralysis files" segments count if callers
  are identifiable.

FIELD SPECIFICATIONS (strict types — no exceptions):

caller_name (string | null):
  Exact name as spoken. If anonymous or unnamed, return null.

country (string | null):
  Country where THE ENCOUNTER TOOK PLACE — not where the caller currently lives.
  Full country name. Default "USA" when a US state is mentioned. null if unknown.

state_or_region (string | null):
  State/region where THE ENCOUNTER TOOK PLACE — not the caller's current residence.
  A caller may say "I live in Texas but this happened when I was visiting Oregon" — use Oregon.
  Full US state name (e.g. "California" not "CA", "Cali", or "SoCal").
  For non-US, use province/region name. null if unknown or withheld.

city (string | null):
  City where THE ENCOUNTER TOOK PLACE — not the caller's current city.
  Only if explicitly mentioned in connection with the event. null otherwise.

call_type (string):
  MUST be one of the exact labels from the ENTITY TAXONOMY below.
  Never invent new labels. Pick the closest match.

call_type_secondary (string | null):
  Second classification from the taxonomy if applicable. null if single type.

description (string):
  1-2 sentence factual summary. No speculation.

date_of_event (string | null):
  ISO-8601 format ONLY. Apply these rules strictly:
  - Exact date known:       "2025-12-02"
  - Only month and year:    "2025-12-01" (first of month)
  - Only year known:        "2025-01-01" (January 1st of that year)
  - Only decade ("the 90s"): "1990-01-01" (first day of decade)
  - "A few years ago" from a 2025 episode: "2022-01-01" (best estimate year)
  - "When I was a kid" with no other clues: null
  - Completely unknown:     null
  NEVER return ranges, approximations, or prose like "mid-90s" or "summer 2020".

time_of_event (string | null):
  24-hour format HH:MM when a specific time is stated or clearly implied:
  - "2 AM" → "02:00"
  - "around 3 in the afternoon" → "15:00"
  - "midnight" → "00:00"
  - "noon" → "12:00"
  - "dusk" → "19:00"
  - "dawn" / "sunrise" → "06:00"
  - "late night" → "23:00"
  - "evening" → "20:00"
  - "early morning" → "05:00"
  - "middle of the night" → "03:00"
  - "pre-dawn" → "04:00"
  - "daytime" with no specifics → null
  - Completely unknown → null

setting (string | null):
  Brief location descriptor (e.g. "bedroom", "highway", "hotel room"). null if unknown.

involves_other_witnesses (boolean):
  MUST be JSON true or false. Never null, never a string.
  If ambiguous, assess whether any other person is described as perceiving the event.
  Default to false if unclear.

caller_emotional_tone (string):
  MUST be exactly one of: "scared", "calm", "nostalgic", "humorous", "matter-of-fact", "emotional"
  No variations, no combinations. Pick the dominant tone. Never null.

derek_commentary (string | null):
  Brief note on Derek's reaction if notable. null if unremarkable.

caller_intro_snippet (string | null):
  The EXACT first 8-15 words the caller says when they begin speaking.
  Copy verbatim from transcript. Used for timestamp lookup. null if not identifiable.

ENTITY TAXONOMY (use these EXACT labels for call_type and call_type_secondary):
{ENTITY_TAXONOMY}

OUTPUT FORMAT:
Return ONLY a valid JSON array of objects. No markdown fences, no backticks,
no explanatory text before or after the array.
"""


# ─────────────────────────────────────────────────────────────────────────────
# POST-LLM SANITIZER — enforces tidy data types regardless of what Claude returns
# ─────────────────────────────────────────────────────────────────────────────
VALID_TONES = {"scared", "calm", "nostalgic", "humorous", "matter-of-fact", "emotional"}

VALID_ENTITY_LABELS = {
    "Bigfoot/Sasquatch","Dogman","Mothman","Skinwalker","Wendigo","Thunderbird",
    "Crawler/Rake","Chupacabra","Mutant Canine/Turner Beast","Lake Monster",
    "Black-Eyed Kids (BEK)","Not Deer","Goatman","Jersey Devil","Unknown Cryptid",
    "Ghost/Apparition","Shadow Person","Hat Man","Old Hag","Poltergeist",
    "Demonic Entity","Angel/Positive Spirit","Doppelganger","Residual Haunting",
    "Intelligent Haunting","Portal/Vortex","Afterlife Communication (ADC)",
    "Sleep Paralysis Entity",
    "UFO/UAP","Black Triangle","Orb/Light","Fireball","Phantom Lights",
    "Abduction","Alien Entity","Satellite/Rocket (Explained)",
    "Missing Time","Glitch in the Matrix","Time Slip","Premonition/Precognition",
    "Telepathy","Psychic Experience","Astral Projection","Aura/Energy",
    "Synchronicity","Men in Black","Phantom Person","Vanishing Object",
    "Dime/Coin from the Dead",
    "Phantom Sound","Phantom Smell","Unexplained Animal Behavior",
    "Electronic Malfunction","Temperature Anomaly",
    "Coincidence","Unidentified Animal","Explained (mundane)","Dream/Vision",
    "Curse/Hex","Fairy/Fae","Humanoid (unclassified)","Other",
}

STATE_ABBREV_TO_FULL = {
    "AL":"Alabama","AK":"Alaska","AZ":"Arizona","AR":"Arkansas","CA":"California",
    "CO":"Colorado","CT":"Connecticut","DE":"Delaware","FL":"Florida","GA":"Georgia",
    "HI":"Hawaii","ID":"Idaho","IL":"Illinois","IN":"Indiana","IA":"Iowa",
    "KS":"Kansas","KY":"Kentucky","LA":"Louisiana","ME":"Maine","MD":"Maryland",
    "MA":"Massachusetts","MI":"Michigan","MN":"Minnesota","MS":"Mississippi",
    "MO":"Missouri","MT":"Montana","NE":"Nebraska","NV":"Nevada","NH":"New Hampshire",
    "NJ":"New Jersey","NM":"New Mexico","NY":"New York","NC":"North Carolina",
    "ND":"North Dakota","OH":"Ohio","OK":"Oklahoma","OR":"Oregon","PA":"Pennsylvania",
    "RI":"Rhode Island","SC":"South Carolina","SD":"South Dakota","TN":"Tennessee",
    "TX":"Texas","UT":"Utah","VT":"Vermont","VA":"Virginia","WA":"Washington",
    "WV":"West Virginia","WI":"Wisconsin","WY":"Wyoming","DC":"District of Columbia",
}

# Common informal names → canonical full state name
STATE_ALIASES = {
    "cali": "California", "socal": "California", "norcal": "California",
    "mass": "Massachusetts", "mich": "Michigan", "minn": "Minnesota",
    "penn": "Pennsylvania", "tenn": "Tennessee", "wisc": "Wisconsin",
    "conn": "Connecticut", "wash": "Washington", "ore": "Oregon",
    "okla": "Oklahoma", "mont": "Montana", "miss": "Mississippi",
    "ala": "Alabama", "ariz": "Arizona", "ark": "Arkansas",
    "colo": "Colorado", "dela": "Delaware", "fla": "Florida",
    "ind": "Indiana", "neb": "Nebraska", "nev": "Nevada",
}

# Time-of-day words → 24h clock
TIME_WORD_MAP = {
    "midnight": "00:00", "middle of the night": "03:00", "pre-dawn": "04:00",
    "predawn": "04:00", "early morning": "05:00", "dawn": "06:00",
    "sunrise": "06:00", "morning": "08:00", "mid-morning": "10:00",
    "noon": "12:00", "midday": "12:00", "afternoon": "14:00",
    "late afternoon": "16:00", "dusk": "19:00", "sunset": "19:00",
    "twilight": "19:00", "evening": "20:00", "night": "21:00",
    "nighttime": "21:00", "late night": "23:00", "late at night": "23:00",
}

NULL_STRINGS = {
    "na", "n/a", "none", "null", "unknown", "unclear", "unspecified",
    "parts unknown", "anonymous", "", "not mentioned", "not specified",
}


def _sanitize_call(call: dict) -> dict:
    """
    Programmatic enforcement of tidy data types on a single call record.
    Runs AFTER Claude returns JSON — catches any inconsistencies the LLM missed.
    """

    # ── Null normalization: coerce sentinel strings to None ──
    for key in ["state_or_region", "city", "country", "call_type_secondary",
                "date_of_event", "time_of_event", "setting", "derek_commentary",
                "caller_intro_snippet"]:
        val = call.get(key)
        if val is None:
            continue
        if isinstance(val, str) and val.strip().lower() in NULL_STRINGS:
            call[key] = None

    # caller_name: coerce "Anonymous" / sentinel strings to null
    name = call.get("caller_name")
    if isinstance(name, str) and name.strip().lower() in NULL_STRINGS:
        call["caller_name"] = None

    # ── State normalization: abbreviations and aliases → full name ──
    state = call.get("state_or_region")
    if isinstance(state, str):
        s = state.strip()
        upper = s.upper()
        lower = s.lower()
        if upper in STATE_ABBREV_TO_FULL:
            call["state_or_region"] = STATE_ABBREV_TO_FULL[upper]
        elif lower in STATE_ALIASES:
            call["state_or_region"] = STATE_ALIASES[lower]
        elif lower in NULL_STRINGS:
            call["state_or_region"] = None
        else:
            # Title-case normalize for consistency ("california" → "California")
            call["state_or_region"] = s.title() if s == s.lower() or s == s.upper() else s

    # ── Country: ensure "USA" when we have a valid US state ──
    if call.get("state_or_region") and call["state_or_region"] in US_REGIONS:
        call["country"] = "USA"

    # ── call_type validation: must be in taxonomy ──
    ct = call.get("call_type") or ""
    if ct not in VALID_ENTITY_LABELS:
        if ct:
            # Try case-insensitive match
            for label in VALID_ENTITY_LABELS:
                if label.lower() == ct.lower():
                    call["call_type"] = label
                    break
            else:
                call["call_type"] = "Other"
        else:
            call["call_type"] = "Other"

    ct2 = call.get("call_type_secondary")
    if ct2 is not None and ct2 not in VALID_ENTITY_LABELS:
        for label in VALID_ENTITY_LABELS:
            if label.lower() == ct2.lower():
                call["call_type_secondary"] = label
                break
        else:
            call["call_type_secondary"] = None

    # ── date_of_event: enforce ISO-8601 or null ──
    date_val = call.get("date_of_event")
    if isinstance(date_val, str):
        date_val = date_val.strip()
        # Already valid YYYY-MM-DD
        if re.match(r"^\d{4}-\d{2}-\d{2}$", date_val):
            pass
        # YYYY only → YYYY-01-01
        elif re.match(r"^\d{4}$", date_val):
            call["date_of_event"] = f"{date_val}-01-01"
        # YYYY-MM → YYYY-MM-01
        elif re.match(r"^\d{4}-\d{2}$", date_val):
            call["date_of_event"] = f"{date_val}-01"
        # Decade references: "1990s", "the 80s", "mid-90s", etc.
        elif m := re.search(r"(\d{4})s", date_val):
            # Full decade like "1990s" or "2000s"
            call["date_of_event"] = f"{m.group(1)}-01-01"
        elif m := re.search(r"(\d{2})s", date_val):
            decade_str = m.group(1)
            century = "19" if int(decade_str) >= 30 else "20"
            call["date_of_event"] = f"{century}{decade_str}-01-01"
        else:
            # Anything else non-conforming → null
            call["date_of_event"] = None

    # ── time_of_event: enforce HH:MM or null ──
    time_val = call.get("time_of_event")
    if isinstance(time_val, str):
        tv = time_val.strip().lower()
        # Check if it's already HH:MM
        if re.match(r"^\d{1,2}:\d{2}$", tv):
            # Normalize to zero-padded HH:MM
            parts = tv.split(":")
            call["time_of_event"] = f"{int(parts[0]):02d}:{parts[1]}"
        # Try word map (exact match)
        elif tv in TIME_WORD_MAP:
            call["time_of_event"] = TIME_WORD_MAP[tv]
        # Parse "2 AM", "3:30 PM", "11pm" etc.
        elif m := re.match(r"(?:around\s+)?(\d{1,2})(?::(\d{2}))?\s*(am|pm|a\.m\.|p\.m\.)", tv):
            h = int(m.group(1))
            mins = m.group(2) or "00"
            ampm = m.group(3).replace(".", "")
            if ampm == "pm" and h != 12:
                h += 12
            elif ampm == "am" and h == 12:
                h = 0
            call["time_of_event"] = f"{h:02d}:{mins}"
        # Check partial matches in TIME_WORD_MAP keys
        else:
            matched = False
            for word, hhmm in TIME_WORD_MAP.items():
                if word in tv:
                    call["time_of_event"] = hhmm
                    matched = True
                    break
            if not matched:
                call["time_of_event"] = None

    # ── involves_other_witnesses: enforce boolean ──
    wit = call.get("involves_other_witnesses")
    if isinstance(wit, bool):
        pass  # already correct
    elif isinstance(wit, str):
        call["involves_other_witnesses"] = wit.strip().lower() in ("true", "yes", "1")
    else:
        call["involves_other_witnesses"] = False

    # ── caller_emotional_tone: enforce enum ──
    tone = call.get("caller_emotional_tone") or ""
    if isinstance(tone, str):
        t = tone.strip().lower().replace("_", "-")
        if t in VALID_TONES:
            call["caller_emotional_tone"] = t
        else:
            # Fuzzy map common variations
            tone_map = {
                "frightened": "scared", "terrified": "scared", "afraid": "scared",
                "anxious": "scared", "nervous": "scared", "uneasy": "scared",
                "relaxed": "calm", "composed": "calm", "neutral": "calm",
                "sentimental": "nostalgic", "wistful": "nostalgic", "reflective": "nostalgic",
                "funny": "humorous", "joking": "humorous", "lighthearted": "humorous",
                "factual": "matter-of-fact", "matter of fact": "matter-of-fact",
                "deadpan": "matter-of-fact", "dry": "matter-of-fact",
                "upset": "emotional", "tearful": "emotional", "crying": "emotional",
                "excited": "emotional", "passionate": "emotional", "moved": "emotional",
            }
            call["caller_emotional_tone"] = tone_map.get(t, "matter-of-fact")
    else:
        call["caller_emotional_tone"] = "matter-of-fact"

    return call


# ─────────────────────────────────────────────────────────────────────────────
# LLM ABSTRACTION — supports Anthropic and OpenAI with the same interface
# ─────────────────────────────────────────────────────────────────────────────
_llm_client = None
_llm_provider = None

def _get_llm_client(provider: str):
    """Lazy-init the API client based on provider string."""
    global _llm_client, _llm_provider
    if _llm_client is not None and _llm_provider == provider:
        return _llm_client

    if provider == "anthropic":
        key = os.getenv("ANTHROPIC_API_KEY", "")
        if not key:
            raise EnvironmentError(
                "ANTHROPIC_API_KEY must be set in .env\n"
                "Get a key at https://console.anthropic.com/"
            )
        from anthropic import Anthropic
        _llm_client = Anthropic(api_key=key)
    elif provider == "openai":
        key = os.getenv("OPENAI_API_KEY", "")
        if not key:
            raise EnvironmentError(
                "OPENAI_API_KEY must be set in .env\n"
                "Get a key at https://platform.openai.com/api-keys"
            )
        from openai import OpenAI
        _llm_client = OpenAI(api_key=key)
    else:
        raise ValueError(f"Unknown provider: {provider}")

    _llm_provider = provider
    return _llm_client


def _llm_call(model: str, system_prompt: str, user_content: str, provider: str) -> str:
    """
    Unified LLM call that works with both Anthropic and OpenAI.
    Returns the raw text response.
    """
    client = _get_llm_client(provider)

    if provider == "anthropic":
        response = client.messages.create(
            model=model,
            max_tokens=8000,
            system=system_prompt,
            messages=[{"role": "user", "content": user_content}],
        )
        return response.content[0].text.strip()

    elif provider == "openai":
        response = client.chat.completions.create(
            model=model,
            max_completion_tokens=8000,
            messages=[
                {"role": "developer", "content": system_prompt},
                {"role": "user", "content": user_content},
            ],
        )
        return response.choices[0].message.content.strip()


# ─────────────────────────────────────────────────────────────────────────────
# VADER SENTIMENT ANALYSIS — adds numeric sentiment scores per caller
# ─────────────────────────────────────────────────────────────────────────────
_vader_analyzer = None

def _get_vader():
    """Lazy-init VADER sentiment analyzer (downloads lexicon once)."""
    global _vader_analyzer
    if _vader_analyzer is not None:
        return _vader_analyzer
    try:
        from nltk.sentiment.vader import SentimentIntensityAnalyzer
        import nltk
        # Ensure the lexicon is downloaded
        try:
            nltk.data.find("sentiment/vader_lexicon.zip")
        except LookupError:
            nltk.download("vader_lexicon", quiet=True)
        _vader_analyzer = SentimentIntensityAnalyzer()
        return _vader_analyzer
    except ImportError:
        log.warning("VADER unavailable (pip install nltk). Sentiment scores will be NA.")
        return None


def _score_vader_sentiment(calls: list[dict]) -> None:
    """
    Runs VADER sentiment analysis on each call's description text.
    Adds four float columns to each call dict:
      sentiment_compound  — overall score, -1.0 (most negative) to +1.0 (most positive)
      sentiment_pos       — proportion of text that is positive, 0.0 to 1.0
      sentiment_neg       — proportion of text that is negative, 0.0 to 1.0
      sentiment_neu       — proportion of text that is neutral, 0.0 to 1.0

    All values are floats rounded to 4 decimal places for tidy data.
    If VADER is unavailable or description is missing, values are set to None.
    """
    analyzer = _get_vader()

    for call in calls:
        desc = (call.get("description") or "").strip()
        if not analyzer or not desc:
            call["sentiment_compound"] = None
            call["sentiment_pos"] = None
            call["sentiment_neg"] = None
            call["sentiment_neu"] = None
            continue

        scores = analyzer.polarity_scores(desc)
        call["sentiment_compound"] = round(scores["compound"], 4)
        call["sentiment_pos"] = round(scores["pos"], 4)
        call["sentiment_neg"] = round(scores["neg"], 4)
        call["sentiment_neu"] = round(scores["neu"], 4)


# ─────────────────────────────────────────────────────────────────────────────
# REPARSE REGIONS — targeted re-extraction of encounter locations only
# ─────────────────────────────────────────────────────────────────────────────
REPARSE_REGIONS_PROMPT = """You are a precise data extractor. You will be given a transcript excerpt from a paranormal podcast
and a description of a specific caller's story. Your ONLY job is to determine WHERE THE ENCOUNTER TOOK PLACE.

CRITICAL DISTINCTION:
- Callers often say where they CURRENTLY LIVE and separately where the ENCOUNTER HAPPENED.
- "I'm calling from Texas" or "I live in Ohio" = caller's RESIDENCE (IGNORE THIS)
- "This happened when I was camping in Montana" or "I was driving through rural Oregon" = ENCOUNTER LOCATION (USE THIS)
- If the caller says "I'm from Maine and this happened in my backyard" — the encounter IS in Maine.
- If the caller only mentions one location and it seems to be where they live AND where it happened, use it.
- If NO location is mentioned in connection with the encounter, return null for all fields.

Return ONLY a JSON object with exactly these fields:
{
  "country": "Full country name or null",
  "state_or_region": "Full US state name (e.g. 'California' not 'CA') or province/region, or null",
  "city": "City name or null"
}

Return ONLY the JSON. No markdown, no explanation."""


def reparse_regions(episodes: list[dict], api_info: dict) -> None:
    """
    Targeted reparse: reads each cached parsed JSON, sends the caller's
    description + transcript context to the LLM asking ONLY for encounter
    location, then patches the cached JSON in place.

    This is much cheaper than a full reparse because:
      - The prompt is tiny (~200 tokens system + ~500 tokens per call)
      - Only location fields are updated; everything else stays intact
      - Already-reparsed episodes are skipped (checks for _regions_reparsed flag)
    """
    model_name = api_info["model"]
    provider = api_info["provider"]
    _get_llm_client(provider)

    parsed_files = sorted(PARSED_DIR.glob("*.json"))
    if not parsed_files:
        log.warning("No parsed JSON files found. Run --step parse first.")
        return

    # Build transcript lookup
    total_calls = 0
    updated_calls = 0
    skipped_eps = 0

    for pf in tqdm(parsed_files, desc="Reparsing regions"):
        calls = json.load(open(pf))
        if not calls:
            continue

        # Skip if already reparsed
        if calls[0].get("_regions_reparsed"):
            skipped_eps += 1
            continue

        # Find matching transcript
        eid = pf.stem
        tp = TRANSCRIPTS_DIR / f"{eid}.txt"
        if not tp.exists():
            continue
        transcript = tp.read_text(encoding="utf-8")
        if transcript.startswith("TRANSCRIPT_UNAVAILABLE"):
            continue

        transcript_clean = _strip_ads(transcript)
        changed = False

        for call in calls:
            total_calls += 1
            desc = call.get("description") or ""
            snippet = call.get("caller_intro_snippet") or ""
            name = call.get("caller_name") or "Anonymous"

            # Build a focused context: caller name + snippet + description
            # Plus a chunk of transcript around the snippet for context
            context = ""
            if snippet:
                idx = transcript_clean.lower().find(snippet.lower()[:50])
                if idx >= 0:
                    # Grab ~1500 chars around the snippet (enough for one caller segment)
                    start = max(0, idx - 200)
                    end = min(len(transcript_clean), idx + 1300)
                    context = transcript_clean[start:end]

            if not context:
                # Fallback: use description only (less reliable but better than nothing)
                context = desc

            user_content = (
                f"Caller: {name}\n"
                f"Description: {desc}\n\n"
                f"TRANSCRIPT EXCERPT:\n{context}"
            )

            try:
                raw = _llm_call(model_name, REPARSE_REGIONS_PROMPT, user_content, provider)
                raw = re.sub(r"^```(?:json)?\s*", "", raw)
                raw = re.sub(r"\s*```$", "", raw)
                loc = json.loads(raw)

                # Patch the call with new location data
                for field in ("country", "state_or_region", "city"):
                    new_val = loc.get(field)
                    if isinstance(new_val, str) and new_val.strip().lower() in NULL_STRINGS:
                        new_val = None
                    call[field] = new_val

                # Re-run location sanitization
                state = call.get("state_or_region")
                if isinstance(state, str):
                    s = state.strip()
                    upper = s.upper()
                    lower = s.lower()
                    if upper in STATE_ABBREV_TO_FULL:
                        call["state_or_region"] = STATE_ABBREV_TO_FULL[upper]
                    elif lower in STATE_ALIASES:
                        call["state_or_region"] = STATE_ALIASES[lower]
                    elif lower in NULL_STRINGS:
                        call["state_or_region"] = None
                    else:
                        call["state_or_region"] = s.title() if s == s.lower() or s == s.upper() else s

                if call.get("state_or_region") and call["state_or_region"] in US_REGIONS:
                    call["country"] = "USA"

                updated_calls += 1
                changed = True

            except Exception as e:
                log.error(f"  Region reparse failed for {name}: {e}")

            time.sleep(1.0)  # Rate limit

        # Mark all calls as reparsed and save
        if changed:
            for call in calls:
                call["_regions_reparsed"] = True
            tmp = pf.with_suffix(".json.tmp")
            with open(tmp, "w") as f:
                json.dump(calls, f, indent=2)
            tmp.replace(pf)

    log.info(f"Region reparse complete: {updated_calls}/{total_calls} calls updated, "
             f"{skipped_eps} episodes already done")


# ─────────────────────────────────────────────────────────────────────────────
# HALLUCINATION VERIFICATION — spot-checks parsed calls against transcript
# ─────────────────────────────────────────────────────────────────────────────
def _verify_calls(calls: list[dict], transcript: str) -> list[dict]:
    """
    Checks each parsed call against the original transcript text to flag
    potential hallucinations. Adds a 'verified' field to each call:
      True  — caller name or intro snippet found in transcript
      False — could not verify; may be hallucinated

    This catches the main failure mode: the LLM inventing callers that don't
    exist in the transcript.
    """
    t_lower = transcript.lower()

    for call in calls:
        evidence = []

        # Check 1: caller_intro_snippet found in transcript
        snippet = (call.get("caller_intro_snippet") or "").strip()
        if snippet and snippet.lower() in t_lower:
            evidence.append("snippet_match")
        elif snippet:
            # Try first 5 words only
            words = snippet.lower().split()[:5]
            if len(words) >= 3 and " ".join(words) in t_lower:
                evidence.append("snippet_partial")

        # Check 2: caller name appears in transcript
        name = (call.get("caller_name") or "").strip()
        if name and name.lower() in t_lower:
            evidence.append("name_found")

        # Check 3: state/region appears in transcript
        state = (call.get("state_or_region") or "").strip()
        if state and state.lower() in t_lower:
            evidence.append("state_found")

        # Check 4: key description words appear (at least 3 non-trivial words)
        desc = (call.get("description") or "").strip()
        if desc:
            desc_words = [w.lower() for w in re.findall(r"\b[a-z]{4,}\b", desc.lower())]
            matches = sum(1 for w in desc_words if w in t_lower)
            if matches >= 3:
                evidence.append("desc_words_match")

        # Verdict: verified if ANY evidence found
        call["verified"] = len(evidence) > 0
        call["verification_evidence"] = ",".join(evidence) if evidence else "none"

    # Log summary
    total = len(calls)
    verified = sum(1 for c in calls if c.get("verified"))
    if total > 0:
        pct = verified / total * 100
        log.info(f"  Verification: {verified}/{total} calls verified ({pct:.0f}%)")
        unverified = [c for c in calls if not c.get("verified")]
        for c in unverified[:3]:  # Show first 3 unverified
            log.warning(f"    ⚠ Unverified: {c.get('caller_name','?')} — {c.get('description','?')[:60]}")

    return calls


def parse_transcripts(episodes: list[dict], api_info: dict) -> list[dict]:
    model_name = api_info["model"]
    provider = api_info["provider"]

    _get_llm_client(provider)  # Fail fast if key is missing
    all_calls = []
    instance_counter = 0

    # Live ETA tracking
    episodes_to_parse = []
    for ep in episodes:
        eid = ep["episode_id"]
        pp = PARSED_DIR / f"{eid}.json"
        tp = TRANSCRIPTS_DIR / f"{eid}.txt"
        if pp.exists():
            continue
        if not tp.exists():
            continue
        text = tp.read_text(encoding="utf-8")
        if text.startswith("TRANSCRIPT_UNAVAILABLE"):
            continue
        episodes_to_parse.append(ep)

    total_to_parse = len(episodes_to_parse)
    parsed_times = []  # rolling window of seconds per episode for ETA

    for ep in tqdm(episodes, desc="Parsing"):
        eid = ep["episode_id"]
        transcript_path = TRANSCRIPTS_DIR / f"{eid}.txt"
        timestamps_path = TRANSCRIPTS_DIR / f"{eid}_timestamps.json"
        parsed_path = PARSED_DIR / f"{eid}.json"

        if parsed_path.exists():
            with open(parsed_path) as f:
                cached = json.load(f)
                for call in cached:
                    instance_counter += 1
                    call["instance"] = instance_counter
                all_calls.extend(cached)
            continue

        if not transcript_path.exists():
            continue
        text = transcript_path.read_text(encoding="utf-8")
        if text.startswith("TRANSCRIPT_UNAVAILABLE"):
            continue

        ep_start = time.time()

        text_clean = _strip_ads(text)
        chunks = _chunk_transcript(text_clean)
        episode_calls = []

        for i, chunk in enumerate(chunks):
            try:
                user_content = (
                    f"Episode: {ep['name']}\n"
                    f"Season: {ep.get('season', 'Unknown')}\n"
                    f"Episode Number: {ep.get('episode', 'Unknown')}\n"
                    f"Release Date: {ep['release_date']}\n\n"
                    f"TRANSCRIPT:\n{chunk}"
                )
                raw = _llm_call(model_name, PARSE_SYSTEM_PROMPT, user_content, provider)
                raw = re.sub(r"^```(?:json)?\s*", "", raw)
                raw = re.sub(r"\s*```$", "", raw)
                calls = json.loads(raw)

                for call in calls:
                    call["season"] = ep.get("season")
                    call["episode"] = ep.get("episode")
                    call["episode_title"] = ep["name"]
                    call["release_date"] = ep["release_date"]
                    _sanitize_call(call)

                # Resolve timestamps
                if timestamps_path.exists():
                    _resolve_timestamps(calls, timestamps_path)

                episode_calls.extend(calls)

            except json.JSONDecodeError as e:
                log.error(f"JSON error for {ep['name'][:50]}: {e}")
            except Exception as e:
                log.error(f"API error for {ep['name'][:50]}: {e}")
                time.sleep(5)
            time.sleep(1.5)

        # Verify calls against original transcript
        _verify_calls(episode_calls, text)

        # Score sentiment with VADER (local, no API cost)
        _score_vader_sentiment(episode_calls)

        for call in episode_calls:
            instance_counter += 1
            call["instance"] = instance_counter

        # Atomic write: tmp then rename so interruption can't leave corrupt cache
        tmp_parsed = parsed_path.with_suffix(".json.tmp")
        with open(tmp_parsed, "w") as f:
            json.dump(episode_calls, f, indent=2)
        tmp_parsed.replace(parsed_path)
        all_calls.extend(episode_calls)

        # Live ETA update
        ep_elapsed = time.time() - ep_start
        parsed_times.append(ep_elapsed)
        # Use rolling average of last 10 episodes
        recent = parsed_times[-10:]
        avg_sec = sum(recent) / len(recent)
        done = len(parsed_times)
        remaining = total_to_parse - done
        eta_sec = remaining * avg_sec
        eta_str = str(timedelta(seconds=int(eta_sec)))
        log.info(f"  [{done}/{total_to_parse}] {ep['name'][:40]} — "
                 f"{len(episode_calls)} calls, {ep_elapsed:.1f}s — ETA: {eta_str}")

    return all_calls


def _resolve_timestamps(calls: list[dict], ts_path: Path) -> None:
    try:
        with open(ts_path) as f:
            segments = json.load(f)
    except Exception:
        return

    full_text = ""
    char_to_time = []
    for seg in segments:
        start_pos = len(full_text)
        full_text += seg["text"] + " "
        char_to_time.append((start_pos, len(full_text), seg["start_hms"]))

    full_lower = full_text.lower()

    for call in calls:
        snippet = (call.get("caller_intro_snippet") or "").lower().strip()
        if not snippet:
            call["call_start_time"] = None
            continue

        idx = full_lower.find(snippet)
        if idx == -1:
            words = snippet.split()[:6]
            idx = full_lower.find(" ".join(words))
        if idx == -1:
            name = (call.get("caller_name") or "").lower()
            if name and name != "anonymous":
                for pat in [f"this is {name}", f"my name is {name}", f"call me {name}"]:
                    idx = full_lower.find(pat)
                    if idx != -1:
                        break

        if idx >= 0:
            for cs, ce, hms in char_to_time:
                if cs <= idx < ce:
                    call["call_start_time"] = hms
                    break
            else:
                call["call_start_time"] = None
        else:
            call["call_start_time"] = None


def _strip_ads(text: str) -> str:
    patterns = [
        r"(?i)let's do the sixty second savings challenge.*?rocketmoney\.com/cancel\.",
        r"(?i)have you ever had an edible.*?lumigummies\.com.*?m a u\.",
        r"(?i)do you find yourself wanting to eat better.*?factormeals\.com.*?show notes\.",
    ]
    for p in patterns:
        text = re.sub(p, "[AD REMOVED]", text, flags=re.DOTALL)
    return text


def _chunk_transcript(text: str, max_chars: int = 75000) -> list[str]:
    if len(text) <= max_chars:
        return [text]
    markers = [
        r"(?=Now (?:folks|gang|working|from|moving|I also))",
        r"(?=(?:Alright|All right),? (?:gang|folks))",
        r"(?=(?:Thank you|Thanks),? .{1,30}(?:for (?:calling|sharing|ringing)))",
    ]
    chunks, remaining = [], text
    while len(remaining) > max_chars:
        best = max_chars
        for m in markers:
            for match in reversed(list(re.finditer(m, remaining[:max_chars + 5000]))):
                if max_chars * 0.5 < match.start() < max_chars:
                    best = match.start()
                    break
        chunks.append(remaining[:best])
        remaining = remaining[best:]
    if remaining:
        chunks.append(remaining)
    return chunks


# ─────────────────────────────────────────────────────────────────────────────
# STEP 4: EXPORT TO TIDY XLSX (same structure as v1 with call_start_time)
# ─────────────────────────────────────────────────────────────────────────────
US_REGIONS = {
    "Connecticut":"Northeast","Maine":"Northeast","Massachusetts":"Northeast",
    "New Hampshire":"Northeast","Rhode Island":"Northeast","Vermont":"Northeast",
    "New Jersey":"Northeast","New York":"Northeast","Pennsylvania":"Northeast",
    "Alabama":"Southeast","Arkansas":"Southeast","Delaware":"Southeast",
    "Florida":"Southeast","Georgia":"Southeast","Kentucky":"Southeast",
    "Louisiana":"Southeast","Maryland":"Southeast","Mississippi":"Southeast",
    "North Carolina":"Southeast","Oklahoma":"Southeast","South Carolina":"Southeast",
    "Tennessee":"Southeast","Texas":"Southeast","Virginia":"Southeast",
    "West Virginia":"Southeast",
    "Illinois":"Midwest","Indiana":"Midwest","Iowa":"Midwest",
    "Kansas":"Midwest","Michigan":"Midwest","Minnesota":"Midwest",
    "Missouri":"Midwest","Nebraska":"Midwest","North Dakota":"Midwest",
    "Ohio":"Midwest","South Dakota":"Midwest","Wisconsin":"Midwest",
    "Alaska":"West","Arizona":"West","California":"West",
    "Colorado":"West","Hawaii":"West","Idaho":"West",
    "Montana":"West","Nevada":"West","New Mexico":"West",
    "Oregon":"West","Utah":"West","Washington":"West","Wyoming":"West",
}


def export_to_xlsx(all_calls: list[dict], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Calls"

    headers = [
        "instance","season","episode","episode_title","release_date","call_start_time",
        "caller_name","country","state_or_region","city","us_census_region",
        "call_type","call_type_secondary","description","date_of_event",
        "time_of_event","setting","involves_other_witnesses",
        "caller_emotional_tone","sentiment_compound","sentiment_pos",
        "sentiment_neg","sentiment_neu","derek_commentary","verified",
    ]

    hf = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    hfill = PatternFill("solid", fgColor="2F5496")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tb = Border(bottom=Side(style="thin"),top=Side(style="thin"),
                left=Side(style="thin"),right=Side(style="thin"))

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = hf, hfill, ha, tb

    for ri, call in enumerate(all_calls, 2):
        state = call.get("state_or_region")
        region = US_REGIONS.get(state, None) if call.get("country") == "USA" and state else None
        wit = call.get("involves_other_witnesses")
        row = [
            call.get("instance"),
            call.get("season"),
            call.get("episode"),
            call.get("episode_title"),
            call.get("release_date"),
            call.get("call_start_time"),
            call.get("caller_name"),
            call.get("country"),
            state,
            call.get("city"),
            region,
            call.get("call_type"),
            call.get("call_type_secondary"),
            call.get("description"),
            call.get("date_of_event"),
            call.get("time_of_event"),
            call.get("setting"),
            wit if isinstance(wit, bool) else None,
            call.get("caller_emotional_tone"),
            call.get("sentiment_compound"),
            call.get("sentiment_pos"),
            call.get("sentiment_neg"),
            call.get("sentiment_neu"),
            call.get("derek_commentary"),
            call.get("verified") if isinstance(call.get("verified"), bool) else None,
        ]
        for ci, v in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value="NA" if v is None else v)
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(vertical="top", wrap_text=True)

    cw = {"A":10,"B":8,"C":8,"D":40,"E":12,"F":14,"G":16,"H":10,"I":18,
          "J":16,"K":14,"L":24,"M":24,"N":60,"O":14,"P":12,"Q":28,"R":12,
          "S":16,"T":12,"U":10,"V":10,"W":10,"X":50,"Y":10}
    for col, w in cw.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:Y{len(all_calls)+1}"

    wb.save(output_path)
    log.info(f"Exported {len(all_calls)} calls to {output_path}")


def _versioned_path(path: Path) -> Path:
    """
    If path already exists, append _v2, _v3, etc. instead of overwriting.
    MAU_Tidy_Data.xlsx → MAU_Tidy_Data_v2.xlsx → MAU_Tidy_Data_v3.xlsx
    """
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    # Strip existing _vN suffix to find the base name
    base = re.sub(r"_v\d+$", "", stem)
    version = 2
    while True:
        candidate = parent / f"{base}_v{version}{suffix}"
        if not candidate.exists():
            return candidate
        version += 1


def reset_data(what: str = "all") -> None:
    """
    Clear cached/generated data so the pipeline can be re-run fresh.
    
    what='parsed'  — delete parsed JSON only (re-run Claude parsing)
    what='all'     — delete everything: episode list, transcripts, parsed, output
    """
    import shutil

    if what in ("parsed", "all"):
        if PARSED_DIR.exists():
            count = len(list(PARSED_DIR.glob("*.json")))
            shutil.rmtree(PARSED_DIR)
            PARSED_DIR.mkdir(parents=True, exist_ok=True)
            log.info(f"Reset: deleted {count} parsed JSON files from {PARSED_DIR}")

    if what == "all":
        for d, label in [(TRANSCRIPTS_DIR, "transcripts"), (AUDIO_DIR, "audio")]:
            if d.exists():
                count = len(list(d.iterdir()))
                shutil.rmtree(d)
                d.mkdir(parents=True, exist_ok=True)
                log.info(f"Reset: deleted {count} {label} files from {d}")

        cache = DATA_DIR / "episode_list.json"
        if cache.exists():
            cache.unlink()
            log.info("Reset: deleted episode_list.json cache")

    log.info(f"Reset complete ({what}). Ready for a fresh run.")


# ─────────────────────────────────────────────────────────────────────────────
# TIME/COST ESTIMATOR — prints before each step runs
# ─────────────────────────────────────────────────────────────────────────────
# Whisper real-time multipliers (wall-clock time = audio_duration / speed)
WHISPER_SPEED = {
    "tiny": 32, "base": 16, "small": 6, "medium": 2, "large": 1,
}

def estimate_work(step: str, episodes: list[dict], api_info: dict) -> None:
    """
    Counts what actually needs doing (skipping cached files) and prints
    a time/cost estimate before the work begins.
    """
    avg_ep_minutes = 65  # fallback average episode length

    lines = []

    # ── Transcribe estimate ──
    if step in ("all", "transcribe"):
        need_transcribe = 0
        for ep in episodes:
            eid = ep["episode_id"]
            tp = TRANSCRIPTS_DIR / f"{eid}.txt"
            if tp.exists() and tp.stat().st_size > 500:
                continue
            if not ep.get("audio_url"):
                continue
            need_transcribe += 1

        if need_transcribe > 0:
            speed = WHISPER_SPEED.get(WHISPER_MODEL, 2)
            total_audio_min = need_transcribe * avg_ep_minutes
            wall_min = total_audio_min / speed
            # Download time: ~55MB per ep at ~5MB/s
            dl_min = (need_transcribe * 55) / (5 * 60)
            total_min = wall_min + dl_min

            if total_min < 60:
                time_str = f"~{total_min:.0f} min"
            elif total_min < 1440:
                time_str = f"~{total_min / 60:.1f} hours"
            else:
                time_str = f"~{total_min / 1440:.1f} days"

            lines.append(f"  Transcribe: {need_transcribe} episodes to process "
                         f"({len(episodes) - need_transcribe} cached) — "
                         f"{time_str} (Whisper {WHISPER_MODEL})")
        else:
            lines.append(f"  Transcribe: all {len(episodes)} episodes already cached — skipping")

    # ── Parse estimate ──
    if step in ("all", "parse"):
        need_parse = 0
        for ep in episodes:
            eid = ep["episode_id"]
            pp = PARSED_DIR / f"{eid}.json"
            if pp.exists():
                continue
            tp = TRANSCRIPTS_DIR / f"{eid}.txt"
            if not tp.exists():
                continue
            text = tp.read_text(encoding="utf-8")
            if text.startswith("TRANSCRIPT_UNAVAILABLE"):
                continue
            need_parse += 1

        if need_parse > 0:
            # Token estimates: ~13K input tokens per episode, ~2K output
            input_tokens = need_parse * 13000
            output_tokens = need_parse * 2000

            # Look up pricing from registry
            in_price = api_info["input"]
            out_price = api_info["output"]
            cost = (input_tokens / 1e6) * in_price + (output_tokens / 1e6) * out_price

            # ~3 sec per episode (API latency + rate limit sleep)
            parse_min = (need_parse * 3) / 60

            lines.append(f"  Parse:      {need_parse} episodes to process "
                         f"({len(episodes) - need_parse} cached) — "
                         f"~{parse_min:.0f} min, ~${cost:.2f} ({api_info['provider']}/{api_info['model']})")
        else:
            lines.append(f"  Parse:      all {len(episodes)} episodes already cached — skipping")

    # ── Export estimate ──
    if step in ("all", "export"):
        lines.append(f"  Export:     ~5 seconds")

    if lines:
        log.info("─── Estimate ───────────────────────────────────")
        for line in lines:
            log.info(line)
        log.info("────────────────────────────────────────────────")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    api_choices = list(MODEL_REGISTRY.keys())
    parser = argparse.ArgumentParser(
        description="MAU Podcast Scraper v2 (RSS + Whisper)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="Examples:\n"
               "  python mau_scraper_v2.py --step all --api haiku --limit 5\n"
               "  python mau_scraper_v2.py --step parse --api o4-mini\n"
               "  python mau_scraper_v2.py --step export\n"
               "  python mau_scraper_v2.py --step reset-parsed\n",
    )
    parser.add_argument("--step", choices=["fetch","transcribe","parse","export","all","demo",
                                           "reset","reset-parsed","reparse-regions"],
                        default="demo")
    parser.add_argument("--api", choices=api_choices, default=DEFAULT_API,
                        help=f"LLM to use for parsing (default: {DEFAULT_API}). "
                             f"Options: {', '.join(api_choices)}")
    parser.add_argument("--output", default=str(OUTPUT_DIR / "MAU_Tidy_Data.xlsx"))
    parser.add_argument("--limit", type=int, default=0,
                        help="Process only first N episodes (for testing). 0 = all.")
    args = parser.parse_args()

    # Resolve API info from registry
    api_info = MODEL_REGISTRY[args.api]
    log.info(f"API: {args.api} → {api_info['provider']}/{api_info['model']}")

    if args.step == "reset":
        reset_data("all")
        return

    if args.step == "reset-parsed":
        reset_data("parsed")
        return

    if args.step == "reparse-regions":
        reparse_regions([], api_info=api_info)
        log.info("Run --step export to regenerate the XLSX with updated locations.")
        return

    if args.step == "demo":
        log.info("Demo mode — see MAU_Tidy_Data.xlsx for sample output")
        return

    if args.step in ("all", "fetch"):
        episodes = fetch_episode_list()
    else:
        cache = DATA_DIR / "episode_list.json"
        if not cache.exists():
            raise FileNotFoundError("Run --step fetch first")
        with open(cache) as f:
            episodes = json.load(f)

    # Apply --limit for testing
    if args.limit > 0:
        episodes = episodes[:args.limit]
        log.info(f"Limited to first {args.limit} episodes (--limit)")

    # Print time/cost estimate before starting
    estimate_work(args.step, episodes, api_info)

    if args.step in ("all", "transcribe"):
        fetch_transcripts(episodes)

    if args.step in ("all", "parse"):
        all_calls = parse_transcripts(episodes, api_info=api_info)

    if args.step in ("all", "export"):
        if "all_calls" not in dir():
            all_calls = []
            inst = 0
            for ep in episodes:
                pp = PARSED_DIR / f"{ep['episode_id']}.json"
                if pp.exists():
                    for c in json.load(open(pp)):
                        inst += 1
                        c["instance"] = inst
                        all_calls.append(c)
        export_to_xlsx(all_calls, _versioned_path(Path(args.output)))

    log.info("Pipeline complete.")


if __name__ == "__main__":
    main()
